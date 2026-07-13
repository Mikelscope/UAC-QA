from flask import Flask, request, jsonify
from flask_cors import CORS 

from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

app = Flask(__name__)
EXCEL_FILE = "Auto_Wrapping_Analytics.xlsx"
def create_workbook():

    if not os.path.exists(EXCEL_FILE):

        wb = Workbook()

        ws = wb.active
        ws.title = "July 2026"

        headers = [
            "QA ID",
            "Date",
            "Batch Time",
            "Auto Wrapper",
            "Product Type"
        ]

        # Weight 1 - Weight 26
        for i in range(1, 27):
            headers.append(f"Weight {i}")

        headers.extend([
            "Average Weight",
            "On Spec Count",
            "On Spec %",
            "Underweight Count",
            "Underweight %",
            "Overweight Count",
            "Overweight %"
        ])

        ws.append(headers)

        wb.save(EXCEL_FILE)

# Allow requests from your frontend
CORS(app)


@app.route("/")
def home():
    return "Auto Wrapping Backend is Running!"


@app.route("/save", methods=["POST"])
def save():

    data = request.get_json()

    # Open workbook
    wb = load_workbook(EXCEL_FILE)

    # Current month sheet
    month_name = datetime.now().strftime("%B %Y")

    # Create sheet if it doesn't exist
    if month_name not in wb.sheetnames:

        ws = wb.create_sheet(title=month_name)

        headers = [
            "QA ID",
            "Date",
            "Time Session",
            "Auto Wrapper",
            "Product Type",

            "Weight 1","Weight 2","Weight 3","Weight 4","Weight 5","Weight 6",
            "Weight 7","Weight 8","Weight 9","Weight 10","Weight 11","Weight 12",
            "Weight 13","Weight 14","Weight 15","Weight 16","Weight 17","Weight 18",
            "Weight 19","Weight 20","Weight 21","Weight 22","Weight 23","Weight 24",
            "Weight 25","Weight 26",

            "Average Weight",

            "On Spec Count",
            "On Spec %",
            "Underweight Count",
            "Underweight %",
            "Overweight Count",
            "Overweight %"
        ]

        ws.append(headers)

    else:
        ws = wb[month_name]

    # Generate QA ID
    next_number = ws.max_row
    qa_id = f"{next_number:04d}"

    print("QA ID:", qa_id)

    print("\n===== DATA RECEIVED =====")

    for key, value in data.items():
        print(key, ":", value)

    print("=========================\n")

    wb.save(EXCEL_FILE)

    return jsonify({
        "success": True,
        "qa_id": qa_id  
    })


create_workbook()
if __name__ == "__main__":
    app.run(debug=True)

    



    